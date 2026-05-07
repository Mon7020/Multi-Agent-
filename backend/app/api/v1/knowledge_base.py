"""Knowledge base API for document management and RAG runtime parameters."""

from __future__ import annotations

import os
import posixpath
import re
import zipfile
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import List
from urllib.parse import unquote
import xml.etree.ElementTree as ET

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from pydantic import BaseModel

from app.api.admin.dependencies import require_admin_user, require_authenticated_user
from app.services.knowledge_admin_service import knowledge_admin_service
from app.services.rag_runtime import get_loaded_rag_tool, get_rag_tool, rag_params_manager
from app.services.settings_admin_service import settings_admin_service

router = APIRouter()

_ALLOWED_DOC_EXTENSIONS = {".txt", ".pdf", ".docx", ".html", ".htm", ".xlsx"}
_MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
_XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_XLSX_DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._skip_depth = 0
        self._chunks: List[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        del attrs
        if tag in {"script", "style"}:
            self._skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style"} and self._skip_depth > 0:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        text = data.strip()
        if text:
            self._chunks.append(text)

    def get_text(self) -> str:
        return "\n".join(self._chunks)


class DocumentInfo(BaseModel):
    id: str
    filename: str
    file_path: str
    file_type: str
    chunk_count: int
    size: int
    upload_time: str
    update_time: str


class DocumentListResponse(BaseModel):
    documents: List[DocumentInfo]
    total: int


class DocumentContentResponse(BaseModel):
    id: str
    filename: str
    content: str
    chunks: List[dict]


class UpdateDocumentRequest(BaseModel):
    content: str


class RAGParams(BaseModel):
    chunk_size: int = 400
    chunk_overlap: int = 50
    top_k: int = 5
    similarity_threshold: float = 0.3
    enable_cache: bool = True
    enable_rerank: bool = True
    enable_hybrid: bool = True
    enable_self_rag: bool = False


class RAGParamsResponse(BaseModel):
    params: RAGParams
    cache_stats: dict
    metrics: dict
    frontend_policy: dict


def get_project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def get_docs_dir() -> Path:
    docs_dir = knowledge_admin_service.docs_dir
    docs_dir.mkdir(parents=True, exist_ok=True)
    return docs_dir


def _validate_filename(filename: str) -> str:
    raw = (filename or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="filename is required")

    safe = os.path.basename(raw)
    if safe != raw:
        raise HTTPException(status_code=400, detail="invalid filename")

    if safe in {".", ".."} or "/" in safe or "\\" in safe or "\x00" in safe:
        raise HTTPException(status_code=400, detail="invalid filename")

    return safe


def _decode_document_id(document_id: str) -> str:
    return _validate_filename(unquote(document_id or ""))


def _resolve_document_path(document_id: str) -> Path:
    docs_dir = get_docs_dir().resolve()
    filename = _decode_document_id(document_id)
    file_path = (docs_dir / filename).resolve()

    if not file_path.is_relative_to(docs_dir):
        raise HTTPException(status_code=400, detail="invalid document id")

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="document not found")

    return file_path


def _read_html_content(file_path: Path) -> str:
    try:
        from bs4 import BeautifulSoup

        soup = BeautifulSoup(file_path.read_text(encoding="utf-8"), "html.parser")
        for tag_name in ("script", "style"):
            for node in soup.find_all(tag_name):
                node.decompose()
        return soup.get_text("\n", strip=True)
    except ImportError:
        parser = _HTMLTextExtractor()
        parser.feed(file_path.read_text(encoding="utf-8"))
        parser.close()
        return parser.get_text()


def _load_xlsx_shared_strings(archive: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    shared_strings: List[str] = []
    for string_item in root.findall(f".//{{{_XLSX_MAIN_NS}}}si"):
        text_parts = [
            text_node.text or ""
            for text_node in string_item.findall(f".//{{{_XLSX_MAIN_NS}}}t")
        ]
        shared_strings.append("".join(text_parts))
    return shared_strings


def _xlsx_column_index(cell_reference: str) -> int:
    match = re.match(r"([A-Z]+)", (cell_reference or "").upper())
    if not match:
        return 0

    column_index = 0
    for letter in match.group(1):
        column_index = (column_index * 26) + (ord(letter) - ord("A") + 1)
    return column_index


def _parse_xlsx_cell_value(cell_node: ET.Element, shared_strings: List[str]):
    cell_type = cell_node.attrib.get("t")

    if cell_type == "inlineStr":
        text_parts = [
            text_node.text or ""
            for text_node in cell_node.findall(f".//{{{_XLSX_MAIN_NS}}}t")
        ]
        return "".join(text_parts)

    value_node = cell_node.find(f"{{{_XLSX_MAIN_NS}}}v")
    if value_node is None or value_node.text is None:
        return None

    raw_value = value_node.text
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (ValueError, IndexError):
            return raw_value
    if cell_type == "b":
        return "TRUE" if raw_value == "1" else "FALSE"

    try:
        numeric_value = float(raw_value)
    except ValueError:
        return raw_value

    if numeric_value.is_integer():
        return int(numeric_value)
    return numeric_value


def _parse_xlsx_sheet_rows(sheet_xml: bytes, shared_strings: List[str]) -> List[List[object]]:
    root = ET.fromstring(sheet_xml)
    rows: List[List[object]] = []
    max_columns = 0

    for row_node in root.findall(f".//{{{_XLSX_MAIN_NS}}}sheetData/{{{_XLSX_MAIN_NS}}}row"):
        row_values = {}
        row_max_column = 0

        for cell_index, cell_node in enumerate(
            row_node.findall(f"{{{_XLSX_MAIN_NS}}}c"),
            start=1,
        ):
            column_index = _xlsx_column_index(cell_node.attrib.get("r", "")) or cell_index
            row_values[column_index] = _parse_xlsx_cell_value(cell_node, shared_strings)
            row_max_column = max(row_max_column, column_index)

        if row_max_column == 0:
            rows.append([])
            continue

        max_columns = max(max_columns, row_max_column)
        rows.append([row_values.get(index) for index in range(1, row_max_column + 1)])

    if max_columns > 0:
        for row in rows:
            if len(row) < max_columns:
                row.extend([None] * (max_columns - len(row)))

    return rows


def _read_xlsx_content_from_archive(file_path: Path) -> str:
    with zipfile.ZipFile(file_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationship_targets = {}

        for relationship in relationships_root.findall(f"{{{_XLSX_REL_NS}}}Relationship"):
            relationship_id = relationship.attrib.get("Id")
            target = relationship.attrib.get("Target")
            if not relationship_id or not target:
                continue
            relationship_targets[relationship_id] = posixpath.normpath(
                posixpath.join("xl", target)
            ).lstrip("/")

        lines: List[str] = []
        shared_strings = _load_xlsx_shared_strings(archive)
        for sheet in workbook_root.findall(f".//{{{_XLSX_MAIN_NS}}}sheet"):
            sheet_name = sheet.attrib.get("name") or "Sheet"
            relationship_id = sheet.attrib.get(f"{{{_XLSX_DOC_REL_NS}}}id")
            sheet_path = relationship_targets.get(relationship_id or "")
            if not sheet_path:
                continue

            lines.append(f"[{sheet_name}]")
            for row in _parse_xlsx_sheet_rows(archive.read(sheet_path), shared_strings):
                values = ["" if cell is None else str(cell).strip() for cell in row]
                if any(values):
                    lines.append(" | ".join(values))

        return "\n".join(lines)


def _read_xlsx_content(file_path: Path) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            lines: List[str] = []
            for sheet in workbook.worksheets:
                lines.append(f"[{sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    values = ["" if cell is None else str(cell).strip() for cell in row]
                    if any(values):
                        lines.append(" | ".join(values))
            return "\n".join(lines)
        finally:
            workbook.close()
    except ImportError:
        return _read_xlsx_content_from_archive(file_path)


def _read_document_content(file_path: Path) -> str:
    suffix = file_path.suffix.lower()

    if suffix == ".txt":
        for encoding in ("utf-8", "utf-8-sig", "gb18030"):
            try:
                return file_path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise HTTPException(status_code=500, detail="failed to decode text document")

    if suffix == ".pdf":
        import pdfplumber

        with pdfplumber.open(str(file_path)) as pdf:
            return "\n".join((page.extract_text() or "") for page in pdf.pages)

    if suffix == ".docx":
        import docx

        document = docx.Document(str(file_path))
        return "\n".join(para.text for para in document.paragraphs)

    if suffix in {".html", ".htm"}:
        return _read_html_content(file_path)

    if suffix == ".xlsx":
        return _read_xlsx_content(file_path)

    raise HTTPException(status_code=400, detail="unsupported document type")


@router.get("/knowledge-base", response_model=DocumentListResponse)
async def list_documents(current_user=Depends(require_authenticated_user())):
    documents = knowledge_admin_service.list_frontend_documents(current_user["role"])
    payload = [
        DocumentInfo(
            id=document["document_id"],
            filename=document["filename"],
            file_path=document["storage_path"],
            file_type=document["file_type"],
            chunk_count=document.get("chunk_count", 0),
            size=document["size"],
            upload_time=document["created_at"],
            update_time=document["updated_at"],
        )
        for document in documents
    ]
    return DocumentListResponse(documents=payload, total=len(payload))


@router.post("/knowledge-base/upload")
async def upload_document(
    file: UploadFile = File(...),
    chunk_size: int | None = Form(None),
    chunk_overlap: int | None = Form(None),
    current_user=Depends(require_admin_user("admin", "super_admin")),
):
    filename = _validate_filename(file.filename or "")
    file_ext = Path(filename).suffix.lower()
    if file_ext not in _ALLOWED_DOC_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"unsupported file type, allowed: {', '.join(sorted(_ALLOWED_DOC_EXTENSIONS))}",
        )

    content = await file.read()
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"file too large (max {_MAX_FILE_SIZE // (1024 * 1024)}MB)")

    docs_dir = get_docs_dir()
    file_path = docs_dir / filename

    if file_path.exists():
        file_path.unlink()

    file_path.write_bytes(content)
    response_payload = None
    index_task = None

    try:
        rag_tool = get_rag_tool()

        actual_chunk_size = chunk_size if chunk_size is not None else rag_params_manager.get_chunk_size()
        actual_chunk_overlap = chunk_overlap if chunk_overlap is not None else rag_params_manager.get_chunk_overlap()

        try:
            rag_tool.delete_documents_by_source(str(file_path))
        except Exception as exc:
            print(f"[WARN] failed deleting old chunks: {exc}")

        documents = rag_tool.load_document(
            str(file_path),
            chunk_size=actual_chunk_size,
            chunk_overlap=actual_chunk_overlap,
        )
        doc_ids = rag_tool.add_documents_to_vector_db(documents)

        response_payload = {
            "success": True,
            "message": f"uploaded '{filename}'",
            "filename": filename,
            "chunk_count": len(doc_ids),
            "chunk_size": actual_chunk_size,
            "chunk_overlap": actual_chunk_overlap,
            "file_size": len(content),
        }
    except Exception as exc:
        if file_path.exists():
            file_path.unlink()
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    finally:
        if file_path.exists():
            record = knowledge_admin_service.update_document_access(
                filename,
                visible_to_frontend=False,
                published=False,
                allowed_roles=["user", "operator", "admin", "super_admin"],
                actor_id=current_user["id"],
            )
            index_task = knowledge_admin_service.enqueue_document_index(
                document_id=record["document_id"],
                version_id=record.get("current_version_id"),
                actor_id=current_user["id"],
            )

    if response_payload is not None and index_task is not None:
        response_payload["index_task_id"] = index_task["task_id"]
        response_payload["index_status"] = index_task["status"]
    return response_payload


@router.get("/knowledge-base/params", response_model=RAGParamsResponse)
async def get_rag_params(current_user=Depends(require_authenticated_user())):
    del current_user
    runtime_params = rag_params_manager.get_params()
    frontend_policy = settings_admin_service.get_frontend_policy()

    params = RAGParams(
        chunk_size=runtime_params["chunk_size"],
        chunk_overlap=runtime_params["chunk_overlap"],
        top_k=runtime_params["top_k"],
        similarity_threshold=runtime_params["similarity_threshold"],
        enable_cache=runtime_params["enable_cache"],
        enable_rerank=runtime_params["enable_rerank"],
        enable_hybrid=runtime_params.get("enable_hybrid", True),
        enable_self_rag=runtime_params["enable_self_rag"],
    )

    try:
        rag_tool = get_loaded_rag_tool()
        if rag_tool is None:
            raise RuntimeError("rag tool not loaded")
        cache_stats = rag_tool.get_cache_stats()
        metrics = rag_tool.get_metrics()
    except Exception:
        cache_stats = {}
        metrics = {}

    return RAGParamsResponse(
        params=params,
        cache_stats=cache_stats,
        metrics=metrics,
        frontend_policy=frontend_policy,
    )


@router.post("/knowledge-base/params")
async def update_rag_params(
    params: RAGParams,
    current_user=Depends(require_admin_user("admin", "super_admin")),
):
    updated = settings_admin_service.update_runtime_params(params.model_dump(), actor_id=current_user["id"])
    return {
        "success": True,
        "message": "runtime params updated",
        "params": updated,
    }


@router.post("/knowledge-base/reload", status_code=status.HTTP_202_ACCEPTED)
async def reload_knowledge_base(current_user=Depends(require_admin_user("admin", "super_admin"))):
    try:
        task = knowledge_admin_service.enqueue_reload(actor_id=current_user["id"])
        return {
            "success": True,
            "task_id": task["task_id"],
            "status": task["status"],
            "message": "knowledge reload queued",
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.post("/knowledge-base/clear-cache")
async def clear_cache(current_user=Depends(require_admin_user("admin", "super_admin"))):
    del current_user
    try:
        rag_tool = get_rag_tool()
        rag_tool.clear_cache()
        return {"success": True, "message": "cache cleared"}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/knowledge-base/cache/health")
async def cache_health(current_user=Depends(require_admin_user("admin", "super_admin"))):
    del current_user
    try:
        rag_tool = get_rag_tool()
        cache = rag_tool.cache

        cache_type = "unknown"
        redis_available = False

        if hasattr(cache, "_available"):
            redis_available = bool(cache._available)
            cache_type = "redis" if redis_available else "memory"
        elif hasattr(cache, "get_stats"):
            cache_type = "memory"

        cache_stats = cache.get_stats() if hasattr(cache, "get_stats") else {}
        rag_metrics = rag_tool.metrics.get_summary() if hasattr(rag_tool, "metrics") else {}

        return {
            "success": True,
            "cache_type": cache_type,
            "redis_available": redis_available,
            "cache_stats": cache_stats,
            "rag_metrics": rag_metrics,
        }
    except Exception as exc:
        return {
            "success": False,
            "error": str(exc),
            "cache_type": "unknown",
            "redis_available": False,
        }


@router.get("/knowledge-base/{document_id}", response_model=DocumentContentResponse)
async def get_document(document_id: str, current_user=Depends(require_authenticated_user())):
    try:
        record = knowledge_admin_service.get_frontend_document(
            document_id=document_id,
            role=current_user["role"],
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="document not found") from exc
    file_path = Path(record["storage_path"])
    content = _read_document_content(file_path)

    return DocumentContentResponse(
        id=record["document_id"],
        filename=record["filename"],
        content=content,
        chunks=[],
    )


@router.put("/knowledge-base/{document_id}")
async def update_document(
    document_id: str,
    request: UpdateDocumentRequest,
    current_user=Depends(require_admin_user("admin", "super_admin")),
):
    del current_user
    file_path = _resolve_document_path(document_id)

    if file_path.suffix.lower() != ".txt":
        raise HTTPException(status_code=400, detail="only .txt documents can be edited")

    try:
        file_path.write_text(request.content, encoding="utf-8")

        rag_tool = get_rag_tool()
        try:
            rag_tool.delete_documents_by_source(str(file_path))
        except Exception as exc:
            print(f"[WARN] failed deleting old chunks during update: {exc}")

        documents = rag_tool.load_document(
            str(file_path),
            chunk_size=rag_params_manager.get_chunk_size(),
            chunk_overlap=rag_params_manager.get_chunk_overlap(),
        )
        doc_ids = rag_tool.add_documents_to_vector_db(documents)

        return {
            "success": True,
            "message": f"document '{file_path.name}' updated",
            "chunk_count": len(doc_ids),
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.delete("/knowledge-base/{document_id}")
async def delete_document(document_id: str, current_user=Depends(require_admin_user("admin", "super_admin"))):
    del current_user
    file_path = _resolve_document_path(document_id)

    try:
        source = str(file_path)
        file_path.unlink()

        rag_tool = get_rag_tool()
        deleted_count = 0
        try:
            deleted_count = rag_tool.delete_documents_by_source(source)
        except Exception as exc:
            print(f"[WARN] failed deleting vector chunks: {exc}")

        return {
            "success": True,
            "message": f"document '{file_path.name}' deleted",
            "deleted_chunks": deleted_count,
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
